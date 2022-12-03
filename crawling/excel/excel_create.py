# openpyxl 설치 - 파이썬에서 엑셀을 쉽게 다룰 수 있도록 도와주는 라이브러리
# pip install openpyxl

import openpyxl

# 1.엑셀 만들기
wb = openpyxl.Workbook()

# 2.워크시트 만들기
ws = wb.create_sheet('게임')

# 3.데이터 추가
ws['A1'] = '번호'
ws['B1'] = '이름'

ws['A2'] = '1'
ws['B2'] = '스타'

#4.엑셀 저장
wb.save('game_data_01.xlsx')
wb.save(r'/Users/injinjeong/jeong_workspace/python_workspace/hellopython/crawling/excel/game_data_02.xlsx')
wb.save('/Users/injinjeong/jeong_workspace/python_workspace/hellopython/crawling/excel/game_data_03.xlsx')
