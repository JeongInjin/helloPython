import openpyxl

fpath = 'game_data_01.xlsx'
# 1.엑셀 불러오기
wb_01 = openpyxl.load_workbook(fpath)
# wb_02 = openpyxl.load_workbook(r'/Users/injinjeong/jeong_workspace/python_workspace/hellopython/crawling/excel/game_data_02.xlsx')
# wb_03 = openpyxl.load_workbook('/Users/injinjeong/jeong_workspace/python_workspace/hellopython/crawling/excel/game_data_03.xlsx')

# 2.엑셀 시트선택
ws_01 = wb_01['게임']

# 3.데이터 수정
ws_01['A3'] = '수정 A3'
ws_01['B3'] = '수정 B3'

# 4.저장
wb_01.save(fpath)